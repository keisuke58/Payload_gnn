#!/usr/bin/env python3
"""Extract clean defect windows + channel hints from the TU Darmstadt industrial
defect-position spreadsheet -> clean_windows.json (committed alongside this script).

The author's software markers (cols "Software marker index Start"/"End") index
directly into the 200 kHz inline scan's 347,977 positions, so they give clean
labels with no clock-sync needed. The "Defect measured?" column localises each
defect to a channel ("CH6", "CH7 - 15 cm too early", "All CHs", or "No").

ch field in the JSON:  >0 = that channel number, -1 = all channels, 0 = unknown.

Run:  LD_LIBRARY_PATH=/home/nishioka/miniconda3/lib \
      /home/nishioka/miniconda3/bin/python3.12 parse_windows.py
Needs the xlsx in $TUD_SANDWICH_DATA (gitignored raw data).
"""
import os, re, json
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.environ.get("TUD_SANDWICH_DATA",
                      os.path.abspath(os.path.join(HERE, "..", "..", "data", "external", "tudarmstadt_sandwich")))
XLSX = os.path.join(DATA, "inline-industrial-prototype-defect-positions.xlsx")

ws = load_workbook(XLSX, read_only=True, data_only=True)['Tabelle1']
rows = list(ws.iter_rows(values_only=True))

clean = []
for r in rows[1:]:                       # col7=marker start, col8=end, col5=designator, col14="Defect measured?"
    if not r:
        continue
    s, e, desg, meas = r[7], r[8], r[5], r[14]
    if isinstance(s, (int, float)) and isinstance(e, (int, float)) and e > s:
        ch = 0                           # unknown
        if isinstance(meas, str):
            m = re.search(r'CH\s*(\d+)', meas)
            if m:
                ch = int(m.group(1))
            elif 'all' in meas.lower():
                ch = -1                  # all channels
        clean.append({"s": int(s), "e": int(e), "desg": str(desg), "ch": ch, "raw_ch": str(meas)})

out = os.path.join(HERE, "clean_windows.json")
json.dump(clean, open(out, "w"), indent=0)
from collections import Counter
print(f"wrote {len(clean)} clean windows -> {out}")
print("channel hints:", Counter(str(w['ch']) for w in clean))
